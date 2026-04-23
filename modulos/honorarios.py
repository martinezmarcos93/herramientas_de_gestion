"""
Módulo H: Honorarios
Controlador de aumentos con índice INDEC y exportación
"""
import json
import os
import requests
from datetime import datetime, date
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False


# ─────────────────────────────────────────────
# DATOS: archivo local de inflación
# ─────────────────────────────────────────────
INFLACION_FILE = os.path.join(os.path.dirname(__file__), "..", "config", "inflacion.json")

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

def _cargar_inflacion() -> dict:
    if os.path.exists(INFLACION_FILE):
        with open(INFLACION_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def _guardar_inflacion(data: dict):
    os.makedirs(os.path.dirname(INFLACION_FILE), exist_ok=True)
    with open(INFLACION_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ─────────────────────────────────────────────
# INDEC: carga del IPC desde API pública
# ─────────────────────────────────────────────
def intentar_cargar_indec() -> dict[str, float]:
    """
    Intenta descargar IPC mensual desde la API de datos.gob.ar
    Retorna {año-mes: porcentaje} ej: {"2025-07": 2.90}
    """
    url = "https://apis.datos.gob.ar/series/api/series/"
    params = {
        "ids": "148.3_INIVELGENERAL_DICI_M_26",  # IPC nivel general mensual
        "limit": 100,
        "sort": "desc",
        "format": "json"
    }
    try:
        r = requests.get(url, params=params, timeout=8)
        r.raise_for_status()
        data = r.json()
        resultado = {}
        for item in data.get("data", []):
            fecha_str, valor = item[0], item[1]
            if valor is not None:
                # fecha viene como "2025-07-01"
                partes = fecha_str[:7]  # "2025-07"
                resultado[partes] = round(float(valor), 2)
        return resultado
    except Exception as e:
        return {}


def sincronizar_indec() -> tuple[int, str]:
    """
    Sincroniza datos INDEC con el archivo local.
    Retorna (cantidad_nuevos, mensaje)
    """
    existentes = _cargar_inflacion()
    nuevos = intentar_cargar_indec()
    if not nuevos:
        return 0, "No se pudo conectar con INDEC. Ingresá los valores manualmente."

    agregados = 0
    for k, v in nuevos.items():
        if k not in existentes:
            existentes[k] = v
            agregados += 1

    _guardar_inflacion(existentes)
    return agregados, f"Se actualizaron {agregados} meses desde INDEC."


def get_inflacion(año_mes: str) -> Optional[float]:
    """Obtiene el índice de un mes puntual. Ej: '2025-07' → 2.20"""
    data = _cargar_inflacion()
    return data.get(año_mes)

def set_inflacion_manual(año_mes: str, valor: float):
    """Permite ingresar manualmente un índice."""
    data = _cargar_inflacion()
    data[año_mes] = round(valor, 2)
    _guardar_inflacion(data)

def get_todos_los_indices() -> dict:
    return _cargar_inflacion()


# ─────────────────────────────────────────────
# CÁLCULO DE TABLA DE HONORARIOS
# ─────────────────────────────────────────────
def _año_mes_range(inicio: str, fin: str) -> list[str]:
    """Genera lista de 'YYYY-MM' entre inicio y fin inclusive."""
    resultado = []
    año, mes = int(inicio[:4]), int(inicio[5:7])
    año_f, mes_f = int(fin[:4]), int(fin[5:7])
    while (año, mes) <= (año_f, mes_f):
        resultado.append(f"{año:04d}-{mes:02d}")
        mes += 1
        if mes > 12:
            mes = 1
            año += 1
    return resultado

def calcular_tabla(
    honorario_base: float,
    mes_base: str,
    meses_congelados: list[str] = None,
    meses_recupero: list[str] = None,
    hasta: str = None
) -> list[dict]:
    """
    Calcula la tabla completa de honorarios mes a mes.

    honorario_base: valor del mes_base (ej: 25475.00)
    mes_base: 'YYYY-MM' (ej: '2025-07')
    meses_congelados: lista de meses donde el usuario decidió no aumentar
    meses_recupero: lista de meses donde se recuperan los índices salteados
    hasta: 'YYYY-MM' hasta dónde calcular (default: mes actual + 6)

    Retorna lista de dicts por mes.
    """
    if meses_congelados is None:
        meses_congelados = []
    if meses_recupero is None:
        meses_recupero = []

    if hasta is None:
        hoy = date.today()
        mes_fin = hoy.month + 6
        año_fin = hoy.year
        if mes_fin > 12:
            mes_fin -= 12
            año_fin += 1
        hasta = f"{año_fin:04d}-{mes_fin:02d}"

    indices = get_todos_los_indices()
    meses = _año_mes_range(mes_base, hasta)

    tabla = []
    honorario_actual = honorario_base
    pendientes_acumular = []  # índices no aplicados

    for i, ym in enumerate(meses):
        año, mes_num = int(ym[:4]), int(ym[5:7])
        nombre_mes = f"{MESES_ES[mes_num]} {año}"
        inflacion = indices.get(ym)

        if i == 0:
            # Mes base: valor fijo
            tabla.append({
                "año_mes": ym,
                "mes_nombre": nombre_mes,
                "inflacion": inflacion,
                "honorario": round(honorario_actual, 2),
                "estado": "base",
                "nota": "Valor base"
            })
            continue

        sin_indice = inflacion is None
        congelado_manual = ym in meses_congelados
        es_recupero = ym in meses_recupero

        if sin_indice or congelado_manual:
            # Guardar índice pendiente si hay
            if inflacion and congelado_manual:
                pendientes_acumular.append(inflacion)

            estado = "sin_indice" if sin_indice else "congelado"
            tabla.append({
                "año_mes": ym,
                "mes_nombre": nombre_mes,
                "inflacion": inflacion,
                "honorario": round(honorario_actual, 2),
                "estado": estado,
                "nota": "Sin dato INDEC" if sin_indice else "Congelado por decisión"
            })

        elif es_recupero and pendientes_acumular:
            # Aplicar este mes + todos los pendientes acumulados
            todos = pendientes_acumular + [inflacion]
            pendientes_acumular = []
            for idx in todos:
                honorario_actual = honorario_actual * (1 + idx / 100)
            tabla.append({
                "año_mes": ym,
                "mes_nombre": nombre_mes,
                "inflacion": inflacion,
                "honorario": round(honorario_actual, 2),
                "estado": "recupero",
                "nota": f"Recupero de {len(todos)-1} mes(es) + actual"
            })
        else:
            # Aplicar normalmente
            honorario_actual = honorario_actual * (1 + inflacion / 100)
            tabla.append({
                "año_mes": ym,
                "mes_nombre": nombre_mes,
                "inflacion": inflacion,
                "honorario": round(honorario_actual, 2),
                "estado": "normal",
                "nota": ""
            })

    return tabla


# ─────────────────────────────────────────────
# EXPORTACIÓN A EXCEL
# ─────────────────────────────────────────────
def exportar_excel(tabla: list[dict], destino: str) -> str:
    if not EXCEL_OK:
        raise ImportError("openpyxl no instalado")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Honorarios"

    # Estilos
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    base_fill   = PatternFill("solid", fgColor="D6E4F0")
    cong_fill   = PatternFill("solid", fgColor="FFF2CC")
    recup_fill  = PatternFill("solid", fgColor="E2EFDA")
    sin_fill    = PatternFill("solid", fgColor="FCE4D6")
    money_fmt   = '#,##0.00'
    pct_fmt     = '0.00"%"'
    thin        = Side(style="thin", color="BBBBBB")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    center      = Alignment(horizontal="center", vertical="center")

    headers = ["Mes", "Inflación mensual", "Honorario", "Estado", "Nota"]
    col_widths = [22, 18, 18, 16, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    estado_fill = {
        "base": base_fill,
        "normal": None,
        "congelado": cong_fill,
        "recupero": recup_fill,
        "sin_indice": sin_fill,
    }

    for row_i, item in enumerate(tabla, 2):
        fill = estado_fill.get(item["estado"])
        vals = [
            item["mes_nombre"],
            item["inflacion"] / 100 if item["inflacion"] else None,
            item["honorario"],
            item["estado"].replace("_", " ").capitalize(),
            item["nota"]
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            if fill:
                cell.fill = fill
            cell.border = border
            cell.alignment = center
            if col == 2 and val is not None:
                cell.number_format = pct_fmt
            if col == 3:
                cell.number_format = money_fmt

    ws.freeze_panes = "A2"
    wb.save(destino)
    return destino
